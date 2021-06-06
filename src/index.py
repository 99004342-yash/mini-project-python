"""
Program for working with xl sheets.
"""

import sys
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


class ExcelOperation:
    """
    ExcelOperation class is used for different Excel operations.

    Parameters:
    name (string): Name of the existing file.
    """

    def __init__(self, name):
        self.name = name
        try:
            self.work_book = load_workbook(name)
        except:
            print(f'File: "{name}" does not exist.')
            sys.exit()

    def get_headers(self, work_sheet):
        """
        get_headers is used for getting the headers of respective file

        Parameters:
        work_sheet : Workbook() class is passed
        """
        return [x.value for x in work_sheet[1]]

    def get_all_ps(self):
        """
        get_all_ps returns a list of all th PS no in the xlsx file.
        """
        ps_no = []
        for work_sheet in self.work_book:
            if not work_sheet.max_row < 2:
                ps_no += [work_sheet[get_column_letter(
                    1) + str(i)].value for i in range(2, work_sheet.max_row)]

        return set(ps_no)

    def get_all_data(self, ps_no):
        """
        Function used for generating the output in `output.xlsx`.
        """
        char = get_column_letter(1)

        work_book_2 = Workbook()

        for work_sheet in self.work_book:
            if work_sheet.max_row < 2:
                continue

            for i in range(2, work_sheet.max_row):
                if work_sheet[char + str(i)].value == ps_no:

                    work_sheet_2 = work_book_2.create_sheet(work_sheet.title)
                    work_sheet_2.append(self.get_headers(work_sheet))
                    work_sheet_2.append([x.value for x in work_sheet[i]])
        work_book_2.save("output.xlsx")


xls = ExcelOperation('input.xlsx')

while True:

    choice = int(input("""
    ## MENU ##

    Select option: 
    1. Show all PS Numbers
    2. Create XL of selected PS Number.

    """))

    if choice == 1:
        for ps_no_xl in xls.get_all_ps():
            print(ps_no_xl)

    elif choice == 2:
        ps_no_input = (int(input(
            """
        Please enter a valid PS No
        """
        )))
        xls.get_all_data(ps_no_input)
        print("""
        Please check the output.xlsx!

        """)

    else:
        print("""
        Wrong choice
        
        """)
