"""
Consists of all the unit tests.
"""
import pytest
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from ExcelOperation import ExcelOperation


def test_nameless_instantiating():
    """
    It raises the exception because file name not provided
    """
    with pytest.raises(Exception):
        ExcelOperation()


def test_successful_instantiating():
    """
    Successful Instantiation.
    """
    xl = ExcelOperation('input.xlsx')

    assert isinstance(xl, ExcelOperation) == True


def test_all_ps_no():
    """
    Check if the PS No's are correctly extracted from the XLSheet
    """
    xl = ExcelOperation('input.xlsx')
    work_book = load_workbook('input.xlsx')

    ps_no = []
    for work_sheet in work_book:
        if not work_sheet.max_row < 2:
            ps_no += [work_sheet[get_column_letter(
                1) + str(i)].value for i in range(2, work_sheet.max_row)]

    assert len(set(ps_no)) == len(xl.get_all_ps())
