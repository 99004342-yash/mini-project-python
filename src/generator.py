from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from random import randrange, uniform

randomPs = randrange(99004340, 99004340 + 100)
header = ["PS", "USN"] + \
    [f"SEM_{i}" for i in range(1, 9)] + [f"HOL_{i}" for i in range(1, 11)]

wb = Workbook()

for i in range(1, 6):

    ws = wb.create_sheet(f"Sheet_{i}")


for ws in wb:

    ws.append(header)

    for i in range(0, 15):
        ws.append([randrange(99004340, 99004340 + 100)] +
                  [f"1NT17IS{randrange(100,200)}"] + [round(uniform(6.4, 9.6), 2)for j in range(1, 9)] + [randrange(0, 6) for j in range(0, 10)])

wb.save("input2.xlsx")
